import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os

def adjuntar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("CSV Files", "*.csv")])
    if archivo:
        entry_archivo.delete(0, tk.END)
        entry_archivo.insert(0, archivo)

def generar_informe():
    ruta_csv = entry_archivo.get()
    if not ruta_csv:
        messagebox.showwarning("Advertencia", "Por favor, adjunta un archivo CSV.")
        return

    try:
        df = pd.read_csv(ruta_csv)
        
        columnas_deseadas = [
            'active', 'name.familyName', 'name.givenName', 'userName',
            'urn:ietf:params:scim:schemas:extension:enterprise:2.0:User:division',
            'urn:ietf:params:scim:schemas:extension:enterprise:2.0:User:employeeNumber',
            'urn:ietf:params:scim:schemas:extension:sap:2.0:User:loginTime', 'emails[0].value'
        ]

        columnas_existentes = [col for col in columnas_deseadas if col in df.columns]
        df_reducido = df[columnas_existentes]

        nuevos_nombres = {
            'active': 'Estado',
            'name.familyName': 'Apellido',
            'name.givenName': 'Nombre',
            'userName': 'Usuario',
            'urn:ietf:params:scim:schemas:extension:enterprise:2.0:User:division': 'Gerencia',
            'urn:ietf:params:scim:schemas:extension:enterprise:2.0:User:employeeNumber': 'Area',
            'urn:ietf:params:scim:schemas:extension:sap:2.0:User:loginTime': 'Hora de conexion',
            'emails[0].value': 'correo'
        }

        df_reducido.rename(columns=nuevos_nombres, inplace=True)

        df_reducido = df_reducido[df_reducido['Estado'] == True]
        df_reducido['Estado'] = 'ACTIVO'
        df_reducido = df_reducido[df_reducido['correo'].str.contains('@mallplaza.com', na=False)]
        df_reducido['Hora de conexion'] = pd.to_datetime(df_reducido['Hora de conexion'], errors='coerce').dt.strftime('%Y-%m-%d %H:%M:%S')

        # Generar nombre del archivo con la fecha de hoy
        fecha_hoy = datetime.today().strftime('%d-%m-%Y')
        nombre_archivo = f"Conexion-Usuarios {fecha_hoy}.xlsx"
        ruta_guardado = os.path.join(r"C:\Users\Asus\Documents\Conexion Usuarios\Julio", nombre_archivo)
        
        df_reducido.to_excel(ruta_guardado, index=False, engine='openpyxl')

        wb = load_workbook(ruta_guardado)
        ws = wb.active

        relleno_verde = PatternFill(start_color="5ccb5f", end_color="5ccb5f", fill_type="solid")
        fuente_blanca = Font(color="FFFFFF")

        for cell in ws[1]:
            cell.fill = relleno_verde
            cell.font = fuente_blanca

        wb.save(ruta_guardado)
        messagebox.showinfo("Éxito", f"Informe generado exitosamente en {ruta_guardado}")

    except Exception as e:
        messagebox.showerror("Error", f"Ocurrió un error al procesar el archivo: {e}")

root = tk.Tk()
root.title("Informe Usuarios SAP BTP")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack(padx=10, pady=10)    

label_archivo = tk.Label(frame, text="Archivo")
label_archivo.grid(row=0, column=0, sticky=tk.W, pady=5)

entry_archivo = tk.Entry(frame, width=50)
entry_archivo.grid(row=0, column=1, pady=5)

button_adjuntar = tk.Button(frame, text="Adjuntar", command=adjuntar_archivo)
button_adjuntar.grid(row=0, column=2, padx=5, pady=5)

button_enviar = tk.Button(frame, text="Enviar", command=generar_informe)
button_enviar.grid(row=1, columnspan=3, pady=10)

root.mainloop()
